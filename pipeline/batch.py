from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from .config import AppConfig
from .generation import generate_drafts
from .models import BatchItem, BatchJob
from .repository import article_payload, create_articles, create_material
from .schemas import MaterialInput, normalize_list, normalize_material


MAX_BATCH_ROWS = 200

FIELD_ALIASES = {
    "title_hint": ["title_hint", "title", "标题", "选题", "标题提示"],
    "raw_content": ["raw_content", "content", "正文", "素材", "素材正文", "内容"],
    "keywords": ["keywords", "关键词", "标签"],
    "target_platforms": ["target_platforms", "platforms", "平台", "目标平台"],
    "image_paths": ["image_paths", "图片", "图片路径"],
}

PLATFORM_ALIASES = {
    "公众号": "official_account",
    "微信": "official_account",
    "微信公众平台": "official_account",
    "小红书": "xiaohongshu",
    "知乎": "zhihu",
    "头条": "toutiao",
    "今日头条": "toutiao",
    "视频号": "shipinhao",
}


def pick(row: dict[str, Any], field: str) -> Any:
    for alias in FIELD_ALIASES[field]:
        if alias in row and row[alias] not in {None, ""}:
            return row[alias]
    return ""


def row_to_material(row: dict[str, Any], row_number: int) -> MaterialInput:
    platforms = normalize_list(pick(row, "target_platforms"))
    normalized_platforms = [PLATFORM_ALIASES.get(platform, platform) for platform in platforms]
    return normalize_material(
        {
            "title_hint": pick(row, "title_hint"),
            "raw_content": pick(row, "raw_content"),
            "keywords": pick(row, "keywords"),
            "target_platforms": normalized_platforms,
            "image_paths": pick(row, "image_paths"),
            "source_type": "batch",
            "source_ref": str(row_number),
        }
    )


def parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX 解析需要安装 openpyxl") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item or "").strip() for item in rows[0]]
    result = []
    for values in rows[1:]:
        result.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    return result


def parse_batch_file(filename: str, content: bytes) -> list[MaterialInput]:
    suffix = Path(filename).suffix.lower()
    rows = parse_xlsx(content) if suffix in {".xlsx", ".xlsm"} else parse_csv(content)
    materials: list[MaterialInput] = []
    for index, row in enumerate(rows[:MAX_BATCH_ROWS], start=2):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        materials.append(row_to_material(row, index))
    return materials


def create_batch_job(session: Session, filename: str, materials: list[MaterialInput]) -> BatchJob:
    job = BatchJob(filename=filename, total_count=len(materials), status="pending")
    session.add(job)
    session.flush()
    for index, material in enumerate(materials, start=1):
        item = BatchItem(
            job=job,
            row_number=index,
            status="pending",
            title_hint=material.title_hint,
            raw_content=material.raw_content,
            input_json=json.dumps(
                {
                    "title_hint": material.title_hint,
                    "raw_content": material.raw_content,
                    "keywords": material.keywords,
                    "target_platforms": material.target_platforms,
                    "image_paths": material.image_paths,
                    "source_type": material.source_type,
                    "source_ref": material.source_ref,
                },
                ensure_ascii=False,
            ),
        )
        session.add(item)
    session.flush()
    return job


def process_batch_job(job_id: int, config: AppConfig, session_factory) -> None:
    with session_factory() as session:
        job = session.get(BatchJob, job_id)
        if not job:
            return
        job.status = "running"
        job.updated_at = datetime.utcnow()
        session.commit()

    with session_factory() as session:
        item_ids = list(
            session.scalars(select(BatchItem.id).where(BatchItem.job_id == job_id).order_by(BatchItem.id)).all()
        )

    for item_id in item_ids:
        with session_factory() as session:
            item = session.get(BatchItem, item_id)
            if not item:
                continue
            item.status = "running"
            item.updated_at = datetime.utcnow()
            session.commit()

        try:
            raw_input = json.loads(item.input_json or "{}")
            material_input = normalize_material(raw_input)
            with session_factory() as session:
                material = create_material(session, material_input)
                source, drafts = generate_drafts(material_input, config)
                articles = create_articles(session, material, drafts)
                item = session.get(BatchItem, item_id)
                if item:
                    item.status = "success"
                    item.result_json = json.dumps(
                        {
                            "source": source,
                            "article_ids": [article.id for article in articles],
                            "articles": [article_payload(article) for article in articles],
                        },
                        ensure_ascii=False,
                    )
                    item.updated_at = datetime.utcnow()
                session.commit()
        except Exception as exc:
            with session_factory() as session:
                item = session.get(BatchItem, item_id)
                if item:
                    item.status = "failed"
                    item.error_message = str(exc)
                    item.updated_at = datetime.utcnow()
                session.commit()

    with session_factory() as session:
        job = session.get(BatchJob, job_id)
        if not job:
            return
        items = list(session.scalars(select(BatchItem).where(BatchItem.job_id == job_id)).all())
        job.success_count = sum(1 for item in items if item.status == "success")
        job.failed_count = sum(1 for item in items if item.status == "failed")
        job.status = "success" if job.failed_count == 0 else "partial_failed"
        if job.success_count == 0 and job.failed_count:
            job.status = "failed"
        job.result_message = f"完成 {job.success_count}/{job.total_count}，失败 {job.failed_count}"
        job.updated_at = datetime.utcnow()
        session.commit()


def batch_job_payload(job: BatchJob, include_items: bool = False) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "id": job.id,
        "filename": job.filename,
        "status": job.status,
        "total_count": job.total_count,
        "success_count": job.success_count,
        "failed_count": job.failed_count,
        "result_message": job.result_message,
        "created_at": job.created_at.isoformat(),
        "updated_at": job.updated_at.isoformat(),
    }
    if include_items:
        payload["items"] = [
            {
                "id": item.id,
                "row_number": item.row_number,
                "status": item.status,
                "title_hint": item.title_hint,
                "error_message": item.error_message,
                "result": json.loads(item.result_json or "{}"),
            }
            for item in job.items
        ]
    return payload
